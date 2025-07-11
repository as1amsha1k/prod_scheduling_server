# this file providdes utility functions for the co module 
from openpyxl import load_workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

from matrix_lookup import matrix_lookup

def get_co(sku1,sku2):
    if sku1 in matrix_lookup and sku2 in matrix_lookup[sku1]:
        print(f"Co value for {sku1} and {sku2}: { matrix_lookup[sku1][sku2]}")

        return matrix_lookup[sku1][sku2]
    else:
        return None
    
    


def build_co_mat(path=''):


# Load workbook and sheet
    wb = load_workbook(path)
    ws = wb['data']

    
    # 1️⃣ Row headers (from B column, B4 down)
    sku1_list = []
    row = 7
    while ws[f'B{row}'].value:
        sku1_list.append(ws[f'B{row}'].value)
        row += 1

    # 2️⃣ SKU2s = row headers: F3 right
    sku2_list = []
    col = 6  # F
    while ws[f'{get_column_letter(col)}3'].value:
        sku2_list.append(ws[f'{get_column_letter(col)}3'].value)
        col += 1

    # 3️⃣ Build the lookup JSON
    lookup_json = {}

    start_data_row = 7
    start_data_col = 6  # Column F

    for i, sku1 in enumerate(sku1_list):
        sku2_values = {}
        for j, sku2 in enumerate(sku2_list):
            value = ws.cell(row=start_data_row + i, column=start_data_col + j).value
            sku2_values[sku2] = value
        lookup_json[sku1] = sku2_values


        # ✅ Now you have your matrix as JSON-compatible structure
    with open("matrix_lookup.py", "w", encoding="utf-8") as f:
        f.write("matrix_lookup = ")
        json.dump(lookup_json, f, indent=2, ensure_ascii=False)


def process_lookup():
    # This function is a placeholder for any additional processing you might want to do
    # with the lookup data after building it.
    expanded_lookup = {}


    for outer_key, inner_dict in matrix_lookup.items():
        if inner_dict is None:
            expanded_lookup[outer_key] = None
            continue

        new_inner_dict = {}

        for inner_key, value in inner_dict.items():
            skus = inner_key.split()
            for sku in skus:
                new_inner_dict[sku] = value

        expanded_lookup[outer_key] = new_inner_dict

    # ✅ Save cleaned version to .py file
    with open("matrix_lookup_cleaned.py", "w", encoding="utf-8") as f:
        f.write("matrix_lookup = ")
        json.dump(expanded_lookup, f, indent=2, ensure_ascii=False)

    print("Expanded lookup:")
    print(json.dumps(expanded_lookup, indent=2, ensure_ascii=False))

if __name__ == "__main__":
    

    sku1 =  "KHCA-12-000022"
    sku2 = "KHCA-12-007770014"
    co_value = get_co(sku1, sku2)
    print(f"Co value for {sku1} and {sku2}: {co_value}")


        

